import os.path

from openpyxl import load_workbook


def get_row_value(rowNo1,cellNo1,rowNo2,cellNo2):
    #path=load_workbook()=r"C:\Users\zubai\PycharmProjects\PytestProjectDemo\sheets\Testdata.xlsx"
    file_path =os.path.join(os.getcwd(),"sheets","Testdata.xlsx")
    path=load_workbook(file_path)
    sheet=path["data"]

    username=sheet.cell(row=rowNo1,column=cellNo1).value
    password=sheet.cell(row=rowNo2,column=cellNo2).value

    return username,password